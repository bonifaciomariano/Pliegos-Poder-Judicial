#!/usr/bin/env python3
"""
Genera pliegos_data.json a partir de las 4 fuentes del Senado:
  - Acuerdos_gestion_Milei.xlsx     (planilla madre, filas TIPO == 'AC')
  - AYUDA_MEMORIA_2026__AC_para_dar_cuenta.csv   (validación "dar_cuenta")
  - Audiencias_publicas_acuerdos.md (fechas reales de audiencias ya realizadas)
  - BOLETIN_DE_REUNIONES_DE_COMISIONES_91_2026.pdf (audiencias programadas)

Uso:
    python3 data_builder.py
    python3 data_builder.py --xlsx otra_planilla.xlsx --out otro.json

Ver README.md en esta carpeta para instrucciones completas.
"""
import argparse
import csv
import json
import re
import sys
import unicodedata
from pathlib import Path

import openpyxl
import pdfplumber

HERE = Path(__file__).resolve().parent

# Audiencias programadas que se cargan a mano cuando llega el aviso como texto
# suelto (mail/mensaje) en vez de un boletín PDF nuevo. Se suman a lo que se
# extraiga del PDF. Si más adelante llega un boletín PDF actualizado que ya
# las incluya, se pueden borrar de acá sin problema.
AUDIENCIAS_PROGRAMADAS_MANUAL = [
    {
        "fecha": "18/08/2026", "hora": "14:00", "salon": "Azul",
        "expedientes": [
            "PE-110/26", "PE-195/26", "PE-205/26", "PE-206/26", "PE-207/26",
            "PE-213/26", "PE-214/26", "PE-215/26", "PE-221/26", "PE-222/26",
            "PE-228/26", "PE-230/26", "PE-232/26", "PE-233/26",
        ],
    },
]

# Fechas de audiencia que el boletín PDF cargado todavía lista como "programada"
# pero que ya ocurrieron (confirmado a mano) -> se descartan para no mostrar
# el badge "programada" sobre algo que ya pasó. Correlato: sumarlas como
# realizadas en Audiencias_publicas_acuerdos.md.
AUDIENCIAS_PROGRAMADAS_VENCIDAS = {"12/08/2026", "13/08/2026", "18/08/2026"}

HYPERLINK_RE = re.compile(r'HYPERLINK\(\s*"([^"]*)"\s*,\s*"([^"]*)"\s*\)', re.IGNORECASE)

PROVINCIAS_CANONICAS = {
    "BUENOS AIRES": "Buenos Aires",
    "CATAMARCA": "Catamarca",
    "CHACO": "Chaco",
    "CHUBUT": "Chubut",
    "CORDOBA": "Córdoba",
    "CORRIENTES": "Corrientes",
    "ENTRE RIOS": "Entre Ríos",
    "FORMOSA": "Formosa",
    "JUJUY": "Jujuy",
    "LA PAMPA": "La Pampa",
    "LA RIOJA": "La Rioja",
    "MENDOZA": "Mendoza",
    "MISIONES": "Misiones",
    "NEUQUEN": "Neuquén",
    "RIO NEGRO": "Río Negro",
    "SALTA": "Salta",
    "SAN JUAN": "San Juan",
    "SAN LUIS": "San Luis",
    "SANTA CRUZ": "Santa Cruz",
    "SANTA FE": "Santa Fe",
    "SANTIAGO DEL ESTERO": "Santiago del Estero",
    "TIERRA DEL FUEGO": "Tierra del Fuego",
    "TUCUMAN": "Tucumán",
}

CARGO_KEYWORDS_JUDICIALES = (
    "JUEZ", "JUEZA", "FISCAL", "DEFENSOR", "DEFENSORA", "VOCAL",
    "CAMARISTA", "CONJUEZ", "CONJUECES", "PROCURADOR", "PROCURADORA",
)

DESIGNAR_RE = re.compile(
    r'DESIGNAR\s+(.+?),?\s+(?:AL|A LA)\s+(?:DR\.|DRA\.|DOCTOR|DOCTORA)\s*([^.\n]+)\.?',
    re.IGNORECASE,
)
NOMBRAMIENTO_RE = re.compile(
    r'NOMBRAMIENTO\s+(?:DEL|DE LA)\s+(.+?),?\s+(?:AL|A LA)\s+(?:DR\.|DRA\.|DOCTOR|DOCTORA)\s*([^.\n]+)\.?',
    re.IGNORECASE,
)
CONJUECES_RE = re.compile(r'DESIGNAR\s+(CONJUECES\s+.+?)\.?\s*$', re.IGNORECASE | re.DOTALL)


def extraer_fecha_valida(valor):
    """Saca una fecha dd/mm/aaaa real de un valor de celda (datetime o texto
    tipo ' 20/05/2026 -'); devuelve None si es un placeholder vacío ('  -')."""
    if not valor:
        return None
    if hasattr(valor, "strftime"):
        return valor.strftime("%d/%m/%Y")
    m = re.search(r'\d{2}/\d{2}/\d{4}', str(valor))
    return m.group(0) if m else None


def strip_accents(s):
    return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")


def hyperlink_parts(cell_value):
    """Extrae (url, texto) de una fórmula =HYPERLINK("url","texto"). Devuelve (None, valor) si no es fórmula."""
    if not cell_value:
        return None, ""
    m = HYPERLINK_RE.search(str(cell_value))
    if m:
        return m.group(1), m.group(2)
    return None, str(cell_value)


def parse_dae_field(raw):
    """Columna F: ' {n sesion} {fecha o vacio} -'. Devuelve (numero_sesion, fecha_dd_mm_aaaa|None)."""
    if not raw:
        return None, None
    text = str(raw).replace("-", " ").strip()
    m = re.match(r'^(\d+)\s*(\d{2}/\d{2}/\d{4})?$', text)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def parse_od_field(raw_text):
    """Texto de la col AI (ORDEN DEL DÍA): ' {n} {anio} ... -'. Devuelve (numero, anio) o (None, None)."""
    if not raw_text:
        return None, None
    text = raw_text.strip()
    m = re.match(r'^(\d+)\s+(\d{4})\b', text)
    if not m:
        return None, None
    return m.group(1), m.group(2)


def parse_situacion_field(raw):
    """Columna SANCIONES/SITUACIÓN EXP: 'AP {fecha} ...' si fue sancionado
    (aprobado), 'RE {fecha} ...' si fue rechazado. Devuelve (estado, fecha)
    con estado en {"AP", "RE", None}."""
    if not raw:
        return None, None
    text = str(raw).strip().upper()
    if text.startswith("AP"):
        m = re.search(r'AP\s+(\d{2}/\d{2}/\d{4})', text)
        return ("AP", m.group(1)) if m else ("AP", None)
    if text.startswith("RE"):
        m = re.search(r'RE\s+(\d{2}/\d{2}/\d{4})', text)
        return ("RE", m.group(1)) if m else ("RE", None)
    return None, None


def extract_cargo_candidato(caratula):
    """Devuelve (cargo, candidato) o (None, None) si no matchea el patrón esperado."""
    text = " ".join((caratula or "").split())  # colapsa saltos de línea / espacios

    if re.search(r'\bCONJUECES\b', text, re.IGNORECASE):
        m = CONJUECES_RE.search(text)
        cargo = m.group(1).strip().rstrip(".") if m else text
        return cargo, "Conjueces (designación plural)"

    m = DESIGNAR_RE.search(text)
    if m:
        return m.group(1).strip().rstrip(","), _limpiar_candidato(m.group(2))

    m = NOMBRAMIENTO_RE.search(text)
    if m:
        return m.group(1).strip().rstrip(","), _limpiar_candidato(m.group(2))

    return None, None


def _limpiar_candidato(nombre):
    """Saca referencias tipo '(REF. P.E. 8/20)' que quedan pegadas al nombre
    cuando el punto de 'REF.' corta antes de tiempo la captura del regex."""
    return re.sub(r'\s*\(.*$', '', nombre).strip()


def label_administrativo(caratula):
    """Título genérico para mensajes de trámite (retiros, asignación de
    salas) que no traen un/a candidato/a puntual para extraer."""
    upper = caratula.upper()
    if "ASIGNA SALAS Y VOCAL" in upper:
        return "Asignación de salas y vocalías"
    if "SOLICITA EL RETIRO" in upper:
        return "Retiro de mensaje(s) de designación"
    return "Mensaje administrativo"


def extract_provincia(cargo_text):
    upper = cargo_text.upper()
    if "CAPITAL FEDERAL" in upper:
        return "CABA"
    m = re.search(
        r'PROV(?:INCIA)?\.?\s+DE(L)?\s+([A-ZÁÉÍÓÚÑ ]+?)(?:,|\.|\s+AL\s|\s+A LA\s|$)',
        upper,
    )
    if m:
        es_del = bool(m.group(1))
        nombre = m.group(2).strip()
        key = strip_accents(nombre)
        if es_del and key not in PROVINCIAS_CANONICAS:
            # "PROV. DEL X" -> el nombre de la provincia es X (Chubut, Chaco, Neuquén...)
            return PROVINCIAS_CANONICAS.get(key, nombre.title())
        return PROVINCIAS_CANONICAS.get(key, nombre.title())
    return "Nacional / Múltiple"


def is_judicial(cargo_text):
    upper = cargo_text.upper()
    return any(kw in upper for kw in CARGO_KEYWORDS_JUDICIALES)


def load_audiencias_realizadas(md_path):
    """PE Nº -> fecha (dd/mm/aaaa) de audiencia pública ya realizada, según el .md."""
    result = {}
    text = md_path.read_text(encoding="utf-8")
    meses = {
        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
        "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
        "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
    }
    for line in text.splitlines():
        line = line.strip().lstrip("*").strip()
        if not line:
            continue
        m_fecha = re.match(r'(\d{1,2}) de (\w+) de (\d{4})', line, re.IGNORECASE)
        if not m_fecha:
            continue
        dia, mes_nombre, anio = m_fecha.groups()
        mes = meses.get(mes_nombre.lower())
        if not mes:
            continue
        fecha = f"{int(dia):02d}/{mes}/{anio}"
        for pe in re.findall(r'(\d+)\s*/\s*(\d{2})\b', line):
            nro, yy = pe
            result[f"PE-{nro}/{yy}"] = fecha
    return result


def load_boletin_programadas(pdf_path):
    """PE Nº -> {fecha, hora, salon} para audiencias públicas de Acuerdos PROGRAMADAS a futuro."""
    result = {}
    full_text = []
    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            full_text.append(page.extract_text() or "")
    text = "\n".join(full_text)

    # Bloques "ACUERDOS - AUDIENCIA PÚBLICA" seguidos de fecha/hora/salón y temario con PE-xxx/yy
    block_re = re.compile(
        r'ACUERDOS\s*-\s*AUDIENCIA\s*P[ÚU]BLICA\s*\n'
        r'.*?(\d{1,2})\s+DE\s+(\w+)[^\n]*?(\d{1,2}:\d{2})\s*H\s*[^\n]*?SAL[ÓO]N\s+([A-ZÁÉÍÓÚÑ\s\-]+)\n'
        r'.*?TEMARIO\s*\n(.*?)(?=\n[A-ZÁÉÍÓÚÑ0-9].{0,60}\n🗓|\Z)',
        re.IGNORECASE | re.DOTALL,
    )
    meses = {
        "enero": "01", "febrero": "02", "marzo": "03", "abril": "04",
        "mayo": "05", "junio": "06", "julio": "07", "agosto": "08",
        "septiembre": "09", "octubre": "10", "noviembre": "11", "diciembre": "12",
    }
    anio_boletin = re.search(r'INFORMACI[ÓO]N AL \d{1,2} DE \w+ DE (\d{4})', text, re.IGNORECASE)
    anio_boletin = anio_boletin.group(1) if anio_boletin else "2026"

    for m in block_re.finditer(text):
        dia, mes_nombre, hora, salon, cuerpo = m.groups()
        mes = meses.get(mes_nombre.lower())
        if not mes:
            continue
        fecha = f"{int(dia):02d}/{mes}/{anio_boletin}"
        salon = " ".join(salon.split()).title()
        for pe_nro, pe_yy in re.findall(r'PE[-\s]*(\d+)\s*/\s*(\d{2})\b', cuerpo):
            expediente = f"PE-{pe_nro}/{pe_yy}"
            result[expediente] = {"fecha": fecha, "hora": hora, "salon": salon}
    return result


def load_dar_cuenta_csv(csv_path):
    """Set de expedientes PE-xxx/yy listados en el ayuda memoria como 'para dar cuenta'."""
    result = set()
    with csv_path.open(encoding="utf-8-sig") as f:
        for row in csv.reader(f):
            if not row:
                continue
            m = re.match(r'PE-(\d+)/(\d{2})', row[0].strip(), re.IGNORECASE)
            if m:
                result.add(f"PE-{m.group(1)}/{m.group(2)}")
    return result


def _fila_a_od(fila):
    """fila: dict con las claves 'Sobre los expedientes'/'Número'/'Periodo'/'Fecha Dictamen'.
    Solo expedientes PE-xxx/yy-AC (acuerdos): la planilla de OD también trae
    proyectos de ley (-PL) y otros que reusan la misma numeración PE-xxx/yy."""
    m = re.match(r'PE-(\d+)/(\d{2})-AC\b', (fila.get("Sobre los expedientes") or "").strip(), re.IGNORECASE)
    if not m:
        return None, None
    expediente = f"PE-{m.group(1)}/{m.group(2)}"
    return expediente, {
        "numero": str(fila.get("Número") or "").strip(),
        "anio": str(fila.get("Periodo") or "").strip(),
        "fecha_dictamen": str(fila.get("Fecha Dictamen") or "").strip() or None,
    }


def _leer_od_csv(path):
    with path.open(encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def _leer_od_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    filas = []
    for row in ws.iter_rows(min_row=2):
        if row[0].value is None:
            continue
        filas.append({h: c.value for h, c in zip(headers, row)})
    return filas


def load_nuevas_od(dir_path):
    """Carpeta con exportaciones sueltas de 'Órdenes del Día' nuevas (.csv o
    .xlsx, columnas Número/Sobre los expedientes/Periodo/Fecha Dictamen) que
    todavía no están en la planilla madre. Se van acumulando ahí a medida que
    llegan — no hace falta borrar las viejas, cada corrida las vuelve a leer
    todas. Devuelve expediente -> {numero, anio, fecha_dictamen}."""
    result = {}
    if not dir_path or not dir_path.exists():
        return result
    archivos = sorted(dir_path.glob("*.csv")) + sorted(dir_path.glob("*.xlsx"))
    for path in archivos:
        filas = _leer_od_csv(path) if path.suffix.lower() == ".csv" else _leer_od_xlsx(path)
        for fila in filas:
            expediente, datos = _fila_a_od(fila)
            if expediente:
                result[expediente] = datos
    return result


def build(xlsx_path, csv_path, md_path, pdf_path, nuevas_od_dir=None):
    wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    ws = wb.active
    headers = [c.value for c in ws[1]]
    col = {name: idx for idx, name in enumerate(headers)}

    audiencias_realizadas = load_audiencias_realizadas(md_path)
    audiencias_programadas = load_boletin_programadas(pdf_path)
    dar_cuenta_csv = load_dar_cuenta_csv(csv_path)
    nuevas_od = load_nuevas_od(nuevas_od_dir)

    # suma las audiencias programadas cargadas a mano (ver AUDIENCIAS_PROGRAMADAS_MANUAL)
    for tanda in AUDIENCIAS_PROGRAMADAS_MANUAL:
        for exp in tanda["expedientes"]:
            audiencias_programadas[exp] = {
                "fecha": tanda["fecha"], "hora": tanda["hora"], "salon": tanda["salon"],
            }
    # descarta las fechas que ya pasaron -sea del boletín o de la carga manual-
    # para no mostrar el badge "programada" sobre algo que ya ocurrió
    # (ver AUDIENCIAS_PROGRAMADAS_VENCIDAS)
    audiencias_programadas = {
        exp: datos for exp, datos in audiencias_programadas.items()
        if datos["fecha"] not in AUDIENCIAS_PROGRAMADAS_VENCIDAS
    }

    pliegos = []
    excluidos_no_judiciales = []
    excluidos_sin_match = []

    for row in ws.iter_rows(min_row=2):
        if row[col["TIPO"]].value != "AC":
            continue

        nro_url, nro_txt = hyperlink_parts(row[col["NRO."]].value)
        anio_url, anio_txt = hyperlink_parts(row[col["AÑO"]].value)
        anio_txt = anio_txt or "2026"
        yy = anio_txt[-2:]
        expediente = f"PE-{nro_txt}/{yy}"
        url = nro_url or f"https://www.senado.gob.ar/parlamentario/comisiones/verExp/{nro_txt}.{yy}/PE/AC"

        caratula = row[col["CARÁTULA"]].value or ""

        # mensajes de trámite (retiro de una designación, asignación de
        # salas/vocalías): se muestran igual que un pliego, pero sin exigirles
        # el patrón "DESIGNAR ... AL/A LA DR." ni la lista de cargos
        # judiciales, porque muchos son plurales o puramente administrativos
        es_administrativo = bool(
            re.search(r'ASIGNA SALAS Y VOCAL[IÍ]AS|SOLICITA EL RETIRO', caratula, re.IGNORECASE)
        )

        cargo, candidato = extract_cargo_candidato(caratula)
        if cargo is None:
            if es_administrativo:
                cargo = " ".join(caratula.split())
                candidato = label_administrativo(caratula)
            else:
                excluidos_sin_match.append({"expediente": expediente, "caratula": caratula.strip()})
                continue

        if not es_administrativo and not is_judicial(cargo):
            excluidos_no_judiciales.append({"expediente": expediente, "cargo": cargo, "caratula": caratula.strip()})
            continue

        provincia = extract_provincia(cargo)

        _, dae_url_txt = hyperlink_parts(row[col["NRO. DAE / DADO CUENTA"]].value)
        dae_raw = dae_url_txt or row[col["NRO. DAE / DADO CUENTA"]].value
        sesion_dae, fecha_dado_cuenta = parse_dae_field(dae_raw)

        fecha_ingreso_dictamen = row[col["FECHA INGRESO DICTAMEN"]].value
        fecha_egreso1 = row[col["FECHA_EGRESO1"]].value
        fecha_mesa_entradas_raw = row[col["MESA DE ENTRADAS"]].value
        fecha_ingreso = None
        if fecha_mesa_entradas_raw:
            m = re.search(r'\d{2}/\d{2}/\d{4}', str(fecha_mesa_entradas_raw))
            fecha_ingreso = m.group(0) if m else None

        _, od_txt = hyperlink_parts(row[col["ORDEN DEL DÍA"]].value)
        od_nro, od_anio = parse_od_field(od_txt)
        if not od_nro and expediente in nuevas_od:
            # la planilla madre todavía no tiene esta OD -> se usa la carga manual
            od_nro = nuevas_od[expediente]["numero"]
            od_anio = nuevas_od[expediente]["anio"]
        # la planilla madre no guarda la fecha de dictamen de la OD; se toma
        # de las exportaciones sueltas en nuevas_od/ (se van acumulando ahí)
        od_fecha_dictamen = nuevas_od.get(expediente, {}).get("fecha_dictamen")

        situacion, fecha_situacion = parse_situacion_field(row[col["SANCIONES/SITUACIÓN EXP"]].value)
        fecha_sancion = fecha_situacion if situacion == "AP" else None
        fecha_rechazo = fecha_situacion if situacion == "RE" else None

        fecha_audiencia = (
            audiencias_realizadas.get(expediente)
            or extraer_fecha_valida(fecha_egreso1)
            or extraer_fecha_valida(fecha_ingreso_dictamen)
        )

        # --- categorización, en orden de prioridad ---
        if not fecha_dado_cuenta:
            categoria = "dar_cuenta"
        elif fecha_sancion:
            categoria = "sancionado"
        elif fecha_rechazo:
            categoria = "rechazado"
        elif od_nro:
            categoria = "con_od"
        else:
            categoria = "sin_audiencia"

        programada = audiencias_programadas.get(expediente)

        pliegos.append({
            "expediente": expediente,
            "url": url,
            "candidato": candidato,
            "cargo": cargo,
            "provincia": provincia,
            "categoria": categoria,
            "timeline": {
                "ingreso": fecha_ingreso,
                "dado_cuenta": {
                    "sesion": sesion_dae,
                    "fecha": fecha_dado_cuenta,
                },
                "audiencia": fecha_audiencia if categoria != "dar_cuenta" else None,
                "orden_del_dia": (
                    {"numero": od_nro, "anio": od_anio, "fecha_dictamen": od_fecha_dictamen}
                    if od_nro else None
                ),
                "sancion": fecha_sancion,
                "rechazo": fecha_rechazo,
            },
            "audiencia_programada": programada,
            # badge transitorio: ya tuvo la audiencia pública pero la comisión
            # todavía no emitió el dictamen/Orden del Día -> se apaga solo apenas
            # el pliego avanza a con_od/sancionado
            "audiencia_sin_dictamen": bool(fecha_audiencia and categoria == "sin_audiencia"),
        })

    # Validación cruzada con el ayuda-memoria de "dar cuenta"
    expedientes_generados = {p["expediente"] for p in pliegos}
    dar_cuenta_generados = {p["expediente"] for p in pliegos if p["categoria"] == "dar_cuenta"}
    faltan_en_planilla = sorted(dar_cuenta_csv - expedientes_generados)
    discrepancias = sorted(dar_cuenta_csv - dar_cuenta_generados - set(faltan_en_planilla))

    provincias = sorted({p["provincia"] for p in pliegos})

    resultado = {
        "generado_al": None,
        "total": len(pliegos),
        "provincias": provincias,
        "pliegos": pliegos,
    }

    log = {
        "excluidos_no_judiciales": excluidos_no_judiciales,
        "excluidos_sin_match_regex": excluidos_sin_match,
        "dar_cuenta_csv_no_encontrados_en_planilla": faltan_en_planilla,
        "dar_cuenta_csv_discrepancias_de_categoria": discrepancias,
        "audiencias_programadas_detectadas": len(audiencias_programadas),
        "audiencias_realizadas_detectadas": len(audiencias_realizadas),
    }

    return resultado, log


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--xlsx", default=HERE / "Acuerdos_gestion_Milei.xlsx")
    parser.add_argument("--csv", default=HERE / "AYUDA_MEMORIA_2026__AC_para_dar_cuenta.csv")
    parser.add_argument("--md", default=HERE / "Audiencias_publicas_acuerdos.md")
    parser.add_argument("--pdf", default=HERE / "BOLETIN_DE_REUNIONES_DE_COMISIONES_91_2026.pdf")
    parser.add_argument("--nuevas-od-dir", default=HERE / "nuevas_od",
                         help="Carpeta con exportaciones .csv/.xlsx de Órdenes del Día nuevas que todavía no están en --xlsx")
    parser.add_argument("--out", default=HERE / "pliegos_data.json")
    args = parser.parse_args()

    xlsx_path, csv_path, md_path, pdf_path, out_path = (
        Path(args.xlsx), Path(args.csv), Path(args.md), Path(args.pdf), Path(args.out)
    )
    nuevas_od_dir = Path(args.nuevas_od_dir) if args.nuevas_od_dir else None
    for p in (xlsx_path, csv_path, md_path, pdf_path):
        if not p.exists():
            sys.exit(f"ERROR: no se encuentra el archivo fuente: {p}")

    resultado, log = build(xlsx_path, csv_path, md_path, pdf_path, nuevas_od_dir)

    import datetime
    resultado["generado_al"] = datetime.date.today().isoformat()

    out_path.write_text(json.dumps(resultado, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"OK: {resultado['total']} pliegos judiciales escritos en {out_path}")
    print()
    print("--- Excluidos: carátula no matchea el patrón de designación ---")
    for item in log["excluidos_sin_match_regex"]:
        print(f"  {item['expediente']}: {item['caratula'][:100]}")
    print()
    print("--- Excluidos: cargo no judicial ---")
    for item in log["excluidos_no_judiciales"]:
        print(f"  {item['expediente']}: {item['cargo'][:100]}")
    print()
    if log["dar_cuenta_csv_no_encontrados_en_planilla"]:
        print("--- AVISO: expedientes del ayuda-memoria (dar cuenta) no encontrados en la planilla ---")
        for exp in log["dar_cuenta_csv_no_encontrados_en_planilla"]:
            print(f"  {exp}")
        print()
    if log["dar_cuenta_csv_discrepancias_de_categoria"]:
        print("--- AVISO: expedientes marcados 'para dar cuenta' en el CSV pero con otra categoría en la planilla ---")
        for exp in log["dar_cuenta_csv_discrepancias_de_categoria"]:
            print(f"  {exp}")
        print()
    print(f"Audiencias realizadas detectadas (md): {log['audiencias_realizadas_detectadas']}")
    print(f"Audiencias programadas detectadas (boletín pdf): {log['audiencias_programadas_detectadas']}")

    sin_fecha_dictamen = [
        p["expediente"] for p in resultado["pliegos"]
        if p["categoria"] == "con_od" and not p["timeline"]["orden_del_dia"]["fecha_dictamen"]
    ]
    if sin_fecha_dictamen:
        print()
        print("--- AVISO: con Orden del Día pero sin fecha de dictamen (no está en nuevas_od/) ---")
        for exp in sin_fecha_dictamen:
            print(f"  {exp}")


if __name__ == "__main__":
    main()
